import csv
from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from ladepark_importer.errors import SourceFormatError
from ladepark_importer.schema import STATION_ID_COLUMN


class TabularSource(Protocol):
    @property
    def headers(self) -> tuple[str, ...]: ...

    def rows(self) -> Iterator[dict[str, object]]: ...


class CsvSource:
    def __init__(self, path: Path) -> None:
        self._path = path
        self._headers, self._header_row, self._delimiter = self._inspect()

    @property
    def headers(self) -> tuple[str, ...]:
        return self._headers

    def _inspect(self) -> tuple[tuple[str, ...], int, str]:
        try:
            with self._path.open("r", encoding="utf-8-sig", newline="") as handle:
                sample = handle.read(16_384)
                delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
                handle.seek(0)
                reader = csv.reader(handle, delimiter=delimiter)
                for row_number, row in enumerate(reader, start=1):
                    headers = tuple(str(value).strip() for value in row)
                    if STATION_ID_COLUMN in headers:
                        return headers, row_number, delimiter
                    if row_number >= 50:
                        break
        except (OSError, UnicodeError, csv.Error, StopIteration) as error:
            raise SourceFormatError(f"CSV kann nicht gelesen werden: {error}") from error
        raise SourceFormatError(
            f"Keine Kopfzeile mit {STATION_ID_COLUMN!r} in den ersten 50 Zeilen gefunden"
        )

    def rows(self) -> Iterator[dict[str, object]]:
        try:
            with self._path.open("r", encoding="utf-8-sig", newline="") as handle:
                source_reader = csv.reader(handle, delimiter=self._delimiter)
                for _ in range(self._header_row - 1):
                    next(source_reader)
                reader = csv.DictReader(handle, delimiter=self._delimiter)
                for row in reader:
                    yield {str(key).strip(): value for key, value in row.items() if key is not None}
        except (OSError, UnicodeError, csv.Error) as error:
            raise SourceFormatError(f"CSV-Daten können nicht gelesen werden: {error}") from error


class XlsxSource:
    def __init__(self, path: Path) -> None:
        self._path = path
        self._sheet_name, self._header_row, self._headers = self._inspect()

    @property
    def headers(self) -> tuple[str, ...]:
        return self._headers

    def _inspect(self) -> tuple[str, int, tuple[str, ...]]:
        try:
            workbook = load_workbook(self._path, read_only=True, data_only=True)
        except (OSError, ValueError) as error:
            raise SourceFormatError(f"XLSX kann nicht gelesen werden: {error}") from error
        try:
            for worksheet in workbook.worksheets:
                for row_number, row in enumerate(
                    worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1
                ):
                    values = tuple("" if value is None else str(value).strip() for value in row)
                    if STATION_ID_COLUMN in values:
                        return worksheet.title, row_number, values
        finally:
            workbook.close()
        raise SourceFormatError(
            f"Keine Kopfzeile mit {STATION_ID_COLUMN!r} in den ersten 50 Zeilen gefunden"
        )

    def rows(self) -> Iterator[dict[str, object]]:
        try:
            workbook = load_workbook(self._path, read_only=True, data_only=True)
        except (OSError, ValueError) as error:
            raise SourceFormatError(f"XLSX kann nicht gelesen werden: {error}") from error
        try:
            worksheet = workbook[self._sheet_name]
            for row in worksheet.iter_rows(min_row=self._header_row + 1, values_only=True):
                values: Sequence[object] = row
                if all(value is None for value in values):
                    continue
                yield dict(zip(self._headers, values, strict=False))
        finally:
            workbook.close()


def open_source(path: Path) -> TabularSource:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return CsvSource(path)
    if suffix == ".xlsx":
        return XlsxSource(path)
    raise SourceFormatError(f"Nicht unterstütztes Dateiformat: {suffix or '<ohne Endung>'}")
